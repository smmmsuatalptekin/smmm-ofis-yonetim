from dotenv import load_dotenv
from pathlib import Path
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

import os
import jwt
import bcrypt
import logging
import calendar as _cal
from datetime import datetime, timezone, timedelta, date
from typing import List, Optional, Any
from fastapi import FastAPI, APIRouter, Request, Response, HTTPException, Depends, UploadFile, File, Form
from fastapi.responses import StreamingResponse, FileResponse
from starlette.middleware.cors import CORSMiddleware
import io
import re
import uuid
import mimetypes
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field
from bson import ObjectId

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

JWT_SECRET = os.environ['JWT_SECRET']
JWT_ALG = "HS256"

app = FastAPI()
api = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("smmm")

# ---------------- helpers ----------------
def now_utc():
    return datetime.now(timezone.utc)

def hash_password(p: str) -> str:
    return bcrypt.hashpw(p.encode(), bcrypt.gensalt()).decode()

def verify_password(p: str, h: str) -> bool:
    try:
        return bcrypt.checkpw(p.encode(), h.encode())
    except Exception:
        return False

def create_access_token(uid, email):
    return jwt.encode({"sub": uid, "email": email, "exp": now_utc()+timedelta(hours=12), "type": "access"}, JWT_SECRET, algorithm=JWT_ALG)

def serialize(doc):
    if not doc:
        return doc
    doc = dict(doc)
    doc['id'] = str(doc.pop('_id'))
    doc.pop('password_hash', None)
    return doc

async def get_current_user(request: Request):
    token = request.cookies.get("access_token")
    if not token:
        ah = request.headers.get("Authorization", "")
        if ah.startswith("Bearer "):
            token = ah[7:]
    if not token:
        raise HTTPException(401, "Not authenticated")
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALG])
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(401, "User not found")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(401, "Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(401, "Invalid token")

async def log_audit(user, action, entity, entity_id=None, detail=None):
    await db.audit_logs.insert_one({
        "user_id": str(user["_id"]), "user_name": user.get("name"), "action": action,
        "entity": entity, "entity_id": entity_id, "detail": detail, "created_at": now_utc().isoformat()
    })

# ---------------- models ----------------
class LoginIn(BaseModel):
    email: str
    password: str

class UserCreate(BaseModel):
    email: str
    password: str
    name: str
    role: str = "personel"

# ---------------- auth ----------------
@api.post("/auth/login")
async def login(body: LoginIn, response: Response):
    email = body.email.lower().strip()
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(body.password, user["password_hash"]):
        raise HTTPException(401, "E-posta veya şifre hatalı")
    token = create_access_token(str(user["_id"]), email)
    response.set_cookie("access_token", token, httponly=True, secure=True, samesite="none", max_age=43200, path="/")
    return serialize(user)

@api.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    return {"ok": True}

@api.get("/auth/me")
async def me(user=Depends(get_current_user)):
    return serialize(user)

# ---------------- users / personel ----------------
@api.get("/users")
async def list_users(user=Depends(get_current_user)):
    docs = await db.users.find().to_list(500)
    return [serialize(d) for d in docs]

@api.post("/users")
async def create_user(body: UserCreate, user=Depends(get_current_user)):
    if user.get("role") not in ("admin", "mali_musavir"):
        raise HTTPException(403, "Yetkiniz yok")
    if await db.users.find_one({"email": body.email.lower()}):
        raise HTTPException(400, "Bu e-posta zaten kayıtlı")
    doc = {"email": body.email.lower(), "password_hash": hash_password(body.password),
           "name": body.name, "role": body.role, "created_at": now_utc().isoformat()}
    r = await db.users.insert_one(doc)
    await log_audit(user, "create", "user", str(r.inserted_id), body.name)
    doc["_id"] = r.inserted_id
    return serialize(doc)

@api.delete("/users/{uid}")
async def delete_user(uid: str, user=Depends(get_current_user)):
    if user.get("role") not in ("admin", "mali_musavir"):
        raise HTTPException(403, "Yetkiniz yok")
    await db.users.delete_one({"_id": ObjectId(uid)})
    return {"ok": True}

# ---------------- clients ----------------
@api.get("/clients")
async def list_clients(user=Depends(get_current_user), q: Optional[str] = None, tur: Optional[str] = None, aktif: Optional[str] = None):
    query = {}
    if q:
        query["$or"] = [{"unvan": {"$regex": q, "$options": "i"}}, {"vkn": {"$regex": q, "$options": "i"}},
                        {"tckn": {"$regex": q, "$options": "i"}}, {"telefon": {"$regex": q, "$options": "i"}},
                        {"yetkili": {"$regex": q, "$options": "i"}}]
    if tur:
        query["sirket_turu"] = tur
    if aktif is not None:
        query["aktif"] = aktif == "true"
    docs = await db.clients.find(query).sort("unvan", 1).to_list(2000)
    return [serialize(d) for d in docs]

@api.get("/clients/{cid}")
async def get_client(cid: str, user=Depends(get_current_user)):
    d = await db.clients.find_one({"_id": ObjectId(cid)})
    if not d:
        raise HTTPException(404, "Mükellef bulunamadı")
    return serialize(d)

@api.post("/clients")
async def create_client(body: dict, user=Depends(get_current_user)):
    body["aktif"] = body.get("aktif", True)
    body["created_at"] = now_utc().isoformat()
    body["sorumlu_personel"] = body.get("sorumlu_personel", user.get("name"))
    r = await db.clients.insert_one(body)
    await log_audit(user, "create", "client", str(r.inserted_id), body.get("unvan"))
    body["_id"] = r.inserted_id
    return serialize(body)

@api.put("/clients/{cid}")
async def update_client(cid: str, body: dict, user=Depends(get_current_user)):
    body.pop("id", None); body.pop("_id", None)
    await db.clients.update_one({"_id": ObjectId(cid)}, {"$set": body})
    await log_audit(user, "update", "client", cid, body.get("unvan"))
    d = await db.clients.find_one({"_id": ObjectId(cid)})
    return serialize(d)

@api.delete("/clients/{cid}")
async def delete_client(cid: str, user=Depends(get_current_user)):
    if user.get("role") not in ("admin", "mali_musavir"):
        raise HTTPException(403, "Yetkiniz yok")
    await db.clients.update_one({"_id": ObjectId(cid)}, {"$set": {"aktif": False, "deleted": True}})
    await log_audit(user, "delete", "client", cid)
    return {"ok": True}

# ---------------- declarations ----------------
DECL_TYPES = ["KDV1", "KDV2", "MUHSGK", "Damga", "Geçici Vergi", "Gelir Vergisi", "Kurumlar", "BA/BS"]
DECL_STATUSES = ["Hazırlanmadı", "Hazırlanıyor", "Kontrol Bekliyor", "Hazır", "Onay Bekliyor", "Gönderildi", "Tahakkuk Alındı", "Ödeme Bekliyor", "Tamamlandı", "Muaf"]

def decl_deadline(dtype, period):
    # period 'YYYY-MM'; simple TR deadlines
    y, m = int(period[:4]), int(period[5:7])
    nm = m + 1 if m < 12 else 1
    ny = y if m < 12 else y + 1
    day = 28 if dtype in ("KDV1", "KDV2") else 26 if dtype == "MUHSGK" else 27
    try:
        return date(ny, nm, day).isoformat()
    except Exception:
        return date(ny, nm, 26).isoformat()

@api.get("/declarations")
async def get_declarations(period: str, user=Depends(get_current_user)):
    clients = await db.clients.find({"aktif": True}).sort("unvan", 1).to_list(2000)
    decls = await db.declarations.find({"period": period}).to_list(20000)
    dmap = {}
    for d in decls:
        dmap[(d["client_id"], d["type"])] = serialize(d)
    result = []
    for c in clients:
        cid = str(c["_id"])
        types = c.get("beyanname_turleri", ["KDV1", "MUHSGK"])
        row = {"client_id": cid, "unvan": c["unvan"], "sirket_turu": c.get("sirket_turu"),
               "sorumlu_personel": c.get("sorumlu_personel"), "cells": {}}
        for t in types:
            key = (cid, t)
            if key in dmap:
                row["cells"][t] = dmap[key]
            else:
                row["cells"][t] = {"type": t, "status": "Hazırlanmadı", "period": period,
                                   "deadline": decl_deadline(t, period), "checklist": {}}
        result.append(row)
    return {"period": period, "types": DECL_TYPES, "statuses": DECL_STATUSES, "rows": result}

@api.post("/declarations/update")
async def update_declaration(body: dict, user=Depends(get_current_user)):
    cid = body["client_id"]; t = body["type"]; period = body["period"]
    upd = {"client_id": cid, "type": t, "period": period, "updated_at": now_utc().isoformat()}
    if "status" in body:
        upd["status"] = body["status"]
    if "checklist" in body:
        upd["checklist"] = body["checklist"]
    if "deadline" not in body:
        upd.setdefault("deadline", decl_deadline(t, period))
    existing = await db.declarations.find_one({"client_id": cid, "type": t, "period": period})
    if existing:
        await db.declarations.update_one({"_id": existing["_id"]}, {"$set": upd})
        doc = await db.declarations.find_one({"_id": existing["_id"]})
    else:
        upd["deadline"] = decl_deadline(t, period)
        upd.setdefault("checklist", {})
        upd.setdefault("status", body.get("status", "Hazırlanıyor"))
        r = await db.declarations.insert_one(upd)
        doc = await db.declarations.find_one({"_id": r.inserted_id})
    await log_audit(user, "update", "declaration", cid, f"{t} {period} -> {body.get('status')}")
    return serialize(doc)

@api.post("/declarations/bulk")
async def bulk_declaration(body: dict, user=Depends(get_current_user)):
    period = body["period"]; t = body["type"]; status = body["status"]; ids = body["client_ids"]
    for cid in ids:
        existing = await db.declarations.find_one({"client_id": cid, "type": t, "period": period})
        if existing:
            await db.declarations.update_one({"_id": existing["_id"]}, {"$set": {"status": status, "updated_at": now_utc().isoformat()}})
        else:
            await db.declarations.insert_one({"client_id": cid, "type": t, "period": period, "status": status,
                                              "deadline": decl_deadline(t, period), "checklist": {}, "updated_at": now_utc().isoformat()})
    await log_audit(user, "bulk_update", "declaration", None, f"{t} {period} -> {status} ({len(ids)})")
    return {"ok": True, "count": len(ids)}

# ---------------- e-defter ----------------
EDEFTER_STEPS = ["kayitlar", "kontrol", "berat_olustur", "berat_yukle", "berat_onay", "arsiv"]

@api.get("/edefter")
async def get_edefter(period: str, user=Depends(get_current_user)):
    clients = await db.clients.find({"aktif": True, "edefter": True}).sort("unvan", 1).to_list(2000)
    recs = await db.edefter.find({"period": period}).to_list(5000)
    rmap = {r["client_id"]: serialize(r) for r in recs}
    rows = []
    for c in clients:
        cid = str(c["_id"])
        r = rmap.get(cid) or {"client_id": cid, "period": period, "steps": {}}
        r["unvan"] = c["unvan"]
        rows.append(r)
    return {"period": period, "steps": EDEFTER_STEPS, "rows": rows}

@api.post("/edefter/update")
async def update_edefter(body: dict, user=Depends(get_current_user)):
    cid = body["client_id"]; period = body["period"]; steps = body["steps"]
    existing = await db.edefter.find_one({"client_id": cid, "period": period})
    if existing:
        await db.edefter.update_one({"_id": existing["_id"]}, {"$set": {"steps": steps, "updated_at": now_utc().isoformat()}})
    else:
        await db.edefter.insert_one({"client_id": cid, "period": period, "steps": steps, "updated_at": now_utc().isoformat()})
    await log_audit(user, "update", "edefter", cid, period)
    return {"ok": True}

# ---------------- cari / transactions ----------------
@api.get("/clients/{cid}/transactions")
async def client_transactions(cid: str, start: Optional[str] = None, end: Optional[str] = None, user=Depends(get_current_user)):
    all_docs = await db.transactions.find({"client_id": cid}).to_list(20000)
    all_txns = [serialize(d) for d in all_docs]
    borc = sum(t["amount"] for t in all_txns if t["type"] == "borc")
    alacak = sum(t["amount"] for t in all_txns if t["type"] == "alacak")
    opening, rows, p_borc, p_alacak, p_end = _statement_calc(all_txns, start, end)
    rows_sorted = sorted(rows, key=lambda t: (t.get("date") or t.get("created_at") or ""), reverse=True)
    return {
        "transactions": rows_sorted, "borc": borc, "alacak": alacak, "bakiye": borc - alacak,
        "filtered": bool(start or end), "opening_balance": opening,
        "period_borc": p_borc, "period_alacak": p_alacak, "period_end_balance": p_end,
    }

@api.post("/transactions")
async def create_transaction(body: dict, user=Depends(get_current_user)):
    body["created_at"] = now_utc().isoformat()
    body.setdefault("date", now_utc().date().isoformat())
    r = await db.transactions.insert_one(body)
    await log_audit(user, "create", "transaction", body["client_id"], f"{body['type']} {body['amount']}")
    body["_id"] = r.inserted_id
    return serialize(body)

@api.put("/transactions/{tid}")
async def update_transaction(tid: str, body: dict, user=Depends(get_current_user)):
    body.pop("id", None); body.pop("_id", None)
    if "amount" in body:
        body["amount"] = float(body["amount"])
    body["updated_at"] = now_utc().isoformat()
    await db.transactions.update_one({"_id": ObjectId(tid)}, {"$set": body})
    await log_audit(user, "update", "transaction", body.get("client_id"), f"{body.get('type')} {body.get('amount')}")
    doc = await db.transactions.find_one({"_id": ObjectId(tid)})
    return serialize(doc)

@api.delete("/transactions/{tid}")
async def delete_transaction(tid: str, user=Depends(get_current_user)):
    await db.transactions.delete_one({"_id": ObjectId(tid)})
    return {"ok": True}

@api.post("/clients/{cid}/opening-balance")
async def set_opening_balance(cid: str, body: dict, user=Depends(get_current_user)):
    client = await db.clients.find_one({"_id": ObjectId(cid)})
    if not client:
        raise HTTPException(404, "Mükellef bulunamadı")
    existing = await db.transactions.find_one({"client_id": cid, "kind": "acilis"})
    if existing and not body.get("force"):
        raise HTTPException(409, "Bu mükellef için zaten bir açılış/devir bakiyesi tanımlı")
    if existing and body.get("force"):
        await db.transactions.delete_one({"_id": existing["_id"]})
    direction = body.get("direction", "borc")  # borc=Müşteri Borçlu, alacak=Müşteri Alacaklı
    amount = float(body.get("amount") or 0)
    if amount <= 0:
        raise HTTPException(400, "Tutar 0'dan büyük olmalıdır")
    doc = {
        "client_id": cid, "type": "borc" if direction == "borc" else "alacak",
        "amount": amount, "kind": "acilis",
        "aciklama": body.get("aciklama") or "Açılış / Devir Bakiyesi",
        "date": body.get("date") or now_utc().date().isoformat(),
        "created_at": now_utc().isoformat(),
    }
    r = await db.transactions.insert_one(doc)
    await log_audit(user, "create", "transaction", cid, f"açılış bakiyesi {direction} {amount}")
    doc["_id"] = r.inserted_id
    return serialize(doc)

# ---------------- cari statement PDF ----------------
def _tl(n):
    s = f"{abs(n):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return ("-" if n < 0 else "") + s + " TL"

def _txn_date(t):
    d = t.get("date")
    if d:
        return str(d)[:10]
    ca = t.get("created_at")
    return str(ca)[:10] if ca else ""

def _tr_date(iso):
    if not iso or len(iso) < 10:
        return iso or "-"
    y, m, d = iso[:4], iso[5:7], iso[8:10]
    return f"{d}.{m}.{y}"

def _txn_label(t):
    if t.get("kind") == "acilis":
        return "Devir / Açılış Bakiyesi"
    return "Borç/Tahakkuk" if t.get("type") == "borc" else "Tahsilat"

# Single source of truth for cari balance across screen, PDF and Excel.
def _statement_calc(all_txns, start=None, end=None):
    ordered = sorted(all_txns, key=lambda t: (_txn_date(t) or "0000-00-00", str(t.get("created_at") or "")))
    opening = 0.0
    if start:
        for t in ordered:
            if _txn_date(t) < start:
                opening += float(t.get("amount") or 0) * (1 if t.get("type") == "borc" else -1)
    running = opening
    rows = []
    p_borc = 0.0
    p_alacak = 0.0
    for t in ordered:
        d = _txn_date(t)
        if start and d < start:
            continue
        if end and d > end:
            continue
        amt = float(t.get("amount") or 0)
        if t.get("type") == "borc":
            running += amt; p_borc += amt
        else:
            running -= amt; p_alacak += amt
        r = dict(t); r["running"] = running
        rows.append(r)
    return opening, rows, p_borc, p_alacak, running

_FONTS_READY = False
def _ensure_fonts():
    global _FONTS_READY
    if _FONTS_READY:
        return
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    base = "/usr/share/fonts/truetype/dejavu"
    pdfmetrics.registerFont(TTFont("DejaVu", f"{base}/DejaVuSans.ttf"))
    pdfmetrics.registerFont(TTFont("DejaVu-Bold", f"{base}/DejaVuSans-Bold.ttf"))
    _FONTS_READY = True

@api.get("/clients/{cid}/statement/pdf")
async def statement_pdf(cid: str, start: Optional[str] = None, end: Optional[str] = None, user=Depends(get_current_user)):
    c = await db.clients.find_one({"_id": ObjectId(cid)})
    if not c:
        raise HTTPException(404, "Mükellef bulunamadı")
    txns = await db.transactions.find({"client_id": cid}).to_list(20000)
    # backward-compatible ordering by real transaction date, then insertion time
    txns.sort(key=lambda t: (_txn_date(t) or "0000-00-00", str(t.get("created_at") or "")))
    if start:
        txns = [t for t in txns if _txn_date(t) >= start]
    if end:
        txns = [t for t in txns if _txn_date(t) <= end]

    _ensure_fonts()
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=18 * mm, bottomMargin=18 * mm,
                            leftMargin=15 * mm, rightMargin=15 * mm, title="Cari Hesap Ekstresi")
    title_st = ParagraphStyle("t", fontName="DejaVu-Bold", fontSize=16, leading=20, alignment=1, spaceAfter=2)
    sub_st = ParagraphStyle("s", fontName="DejaVu", fontSize=9, leading=13, alignment=1, textColor=colors.HexColor("#64748b"))
    lbl_st = ParagraphStyle("l", fontName="DejaVu", fontSize=9.5, leading=15)
    cell = ParagraphStyle("c", fontName="DejaVu", fontSize=8.5, leading=11)

    elems = [Paragraph("CARİ HESAP EKSTRESİ", title_st), Spacer(1, 4)]
    unvan = c.get("unvan", "-")
    vkn = c.get("vkn") or c.get("tckn") or "-"
    if start or end:
        aralik = f"{_tr_date(start) if start else 'Başlangıç'} - {_tr_date(end) if end else 'Bugün'}"
    else:
        aralik = "Tüm Hareketler"
    elems.append(Paragraph(f"<b>{unvan}</b>", ParagraphStyle("f", fontName="DejaVu-Bold", fontSize=11, leading=15, alignment=1)))
    elems.append(Paragraph(f"VKN/TCKN: {vkn}", sub_st))
    elems.append(Paragraph(f"Rapor Aralığı: {aralik}", sub_st))
    elems.append(Paragraph(f"Oluşturma Tarihi: {_tr_date(now_utc().date().isoformat())}", sub_st))
    elems.append(Spacer(1, 10))

    header = ["Tarih", "İşlem Türü", "Açıklama", "Borç", "Alacak", "Bakiye"]
    data = [header]
    if start:
        data.append([
            Paragraph(_tr_date(start), cell),
            Paragraph("Dönem Başı Devir", cell),
            Paragraph("Önceki dönemden devreden bakiye", cell),
            Paragraph("-", cell), Paragraph("-", cell),
            Paragraph(_tl(opening), cell),
        ])
    for t in rows:
        is_borc = t.get("type") == "borc"
        amt = float(t.get("amount") or 0)
        data.append([
            Paragraph(_tr_date(_txn_date(t)), cell),
            Paragraph(_txn_label(t), cell),
            Paragraph((t.get("aciklama") or "-"), cell),
            Paragraph(_tl(amt) if is_borc else "-", cell),
            Paragraph(_tl(amt) if not is_borc else "-", cell),
            Paragraph(_tl(t["running"]), cell),
        ])
    if len(rows) == 0 and not start:
        data.append([Paragraph("Bu aralıkta hareket bulunmamaktadır.", cell), "", "", "", "", ""])

    tbl = Table(data, colWidths=[22 * mm, 26 * mm, 58 * mm, 24 * mm, 24 * mm, 26 * mm], repeatRows=1)
    tbl.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "DejaVu"),
        ("FONTNAME", (0, 0), (-1, 0), "DejaVu-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8.5),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (3, 0), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("LINEBELOW", (0, 0), (-1, -1), 0.4, colors.HexColor("#e2e8f0")),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]))
    elems.append(tbl)
    elems.append(Spacer(1, 14))

    bakiye = p_end
    srows = []
    if start:
        srows.append(["Dönem Başı Bakiyesi", _tl(opening)])
    srows.append([("Dönem Borç Toplamı" if start else "Toplam Borç"), _tl(p_borc)])
    srows.append([("Dönem Alacak Toplamı" if start else "Toplam Alacak"), _tl(p_alacak)])
    srows.append([("Dönem Sonu Bakiyesi" if start else "Güncel Bakiye"), _tl(bakiye)])
    last = len(srows) - 1
    summary = Table(srows, colWidths=[48 * mm, 40 * mm], hAlign="RIGHT")
    summary.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "DejaVu"),
        ("FONTNAME", (0, last), (-1, last), "DejaVu-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LINEABOVE", (0, last), (-1, last), 0.8, colors.HexColor("#0f172a")),
        ("TEXTCOLOR", (1, last), (1, last), colors.HexColor("#e11d48") if bakiye > 0 else colors.HexColor("#059669")),
    ]))
    elems.append(summary)
    elems.append(Spacer(1, 18))
    elems.append(Paragraph("BU EKSTRE ALPTEKİN MALİ MÜŞAVİRLİK TARAFINDAN OLUŞTURULMUŞTUR",
                           ParagraphStyle("ft", fontName="DejaVu", fontSize=7.5, textColor=colors.HexColor("#94a3b8"), alignment=1)))

    doc.build(elems)
    buf.seek(0)
    safe = re.sub(r"[^A-Za-z0-9]+", "_", unvan).strip("_") or "Cari"
    rng = f"{(start or 'tum')}_{(end or 'guncel')}"
    fname = f"{safe}_Cari_Ekstre_{rng}.pdf"
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="{fname}"'})

@api.get("/clients/{cid}/statement/xlsx")
async def statement_xlsx(cid: str, start: Optional[str] = None, end: Optional[str] = None, user=Depends(get_current_user)):
    c = await db.clients.find_one({"_id": ObjectId(cid)})
    if not c:
        raise HTTPException(404, "Mükellef bulunamadı")
    all_docs = await db.transactions.find({"client_id": cid}).to_list(20000)
    all_txns = [serialize(d) for d in all_docs]
    opening, rows, p_borc, p_alacak, p_end = _statement_calc(all_txns, start, end)

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from datetime import date as _date
    TL = '#,##0.00" TL"'

    def _to_date(s):
        try:
            return _date(int(s[:4]), int(s[5:7]), int(s[8:10]))
        except Exception:
            return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Cari Ekstre"
    ws["A1"] = "CARİ HESAP EKSTRESİ"; ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = c.get("unvan", "-"); ws["A2"].font = Font(bold=True, size=11)
    ws["A3"] = f"VKN/TCKN: {c.get('vkn') or c.get('tckn') or '-'}"
    aralik = "Tüm Hareketler" if not (start or end) else f"{_tr_date(start) if start else 'Başlangıç'} - {_tr_date(end) if end else 'Bugün'}"
    ws["A4"] = f"Rapor Aralığı: {aralik}"
    ws["A5"] = f"Oluşturma Tarihi: {_tr_date(now_utc().date().isoformat())}"

    hrow = 7
    for i, h in enumerate(["Tarih", "İşlem Türü", "Açıklama", "Borç", "Alacak", "Bakiye"]):
        cc = ws.cell(row=hrow, column=i + 1, value=h)
        cc.font = Font(bold=True, color="FFFFFF")
        cc.fill = PatternFill("solid", fgColor="0F172A")
        cc.alignment = Alignment(horizontal="center")
    r = hrow + 1
    if start:
        dc = ws.cell(row=r, column=1, value=_to_date(start)); dc.number_format = "DD.MM.YYYY"
        ws.cell(row=r, column=2, value="Dönem Başı Devir")
        ws.cell(row=r, column=3, value="Önceki dönemden devreden bakiye")
        bc = ws.cell(row=r, column=6, value=opening); bc.number_format = TL
        r += 1
    for t in rows:
        is_borc = t.get("type") == "borc"
        amt = float(t.get("amount") or 0)
        dc = ws.cell(row=r, column=1, value=_to_date(_txn_date(t))); dc.number_format = "DD.MM.YYYY"
        ws.cell(row=r, column=2, value=_txn_label(t))
        ws.cell(row=r, column=3, value=t.get("aciklama") or "")
        if is_borc:
            vc = ws.cell(row=r, column=4, value=amt); vc.number_format = TL
        else:
            vc = ws.cell(row=r, column=5, value=amt); vc.number_format = TL
        rc = ws.cell(row=r, column=6, value=t["running"]); rc.number_format = TL
        r += 1
    data_last = r - 1
    r += 1
    summary = []
    if start:
        summary.append(("Dönem Başı Bakiyesi", opening))
    summary.append(("Dönem Borç Toplamı" if start else "Toplam Borç", p_borc))
    summary.append(("Dönem Alacak Toplamı" if start else "Toplam Alacak", p_alacak))
    summary.append(("Dönem Sonu Bakiyesi" if start else "Güncel Bakiye", p_end))
    for lbl, val in summary:
        lc = ws.cell(row=r, column=5, value=lbl); lc.font = Font(bold=True)
        vc = ws.cell(row=r, column=6, value=val); vc.number_format = TL; vc.font = Font(bold=True)
        r += 1

    for i, w in enumerate([14, 22, 46, 16, 16, 18]):
        ws.column_dimensions[chr(65 + i)].width = w
    ws.freeze_panes = f"A{hrow + 1}"
    ws.auto_filter.ref = f"A{hrow}:F{max(hrow, data_last)}"

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    safe = re.sub(r"[^A-Za-z0-9]+", "_", c.get("unvan", "Cari")).strip("_") or "Cari"
    rng = f"{(start or 'tum')}_{(end or 'guncel')}"
    fname = f"{safe}_Cari_Ekstre_{rng}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f'attachment; filename="{fname}"'})

# ---------------- documents (Mükellef Evrak Yönetimi) ----------------
UPLOAD_DIR = os.environ.get("UPLOAD_DIR", str(ROOT_DIR / "uploads"))
MAX_UPLOAD_MB = int(os.environ.get("MAX_UPLOAD_MB", "25"))
ALLOWED_DOC_EXT = {".pdf", ".jpg", ".jpeg", ".png", ".xls", ".xlsx", ".doc", ".docx"}

def _doc_serialize(d):
    d = dict(d)
    d["id"] = str(d.pop("_id"))
    return d

@api.post("/clients/{cid}/documents")
async def upload_document(
    cid: str,
    file: UploadFile = File(...),
    title: str = Form(...),
    category: str = Form("Diğer"),
    subcategory: str = Form(""),
    document_date: str = Form(""),
    period: str = Form(""),
    description: str = Form(""),
    tags: str = Form(""),
    expiry_date: str = Form(""),
    parent_document_id: str = Form(""),
    user=Depends(get_current_user),
):
    client = await db.clients.find_one({"_id": ObjectId(cid)})
    if not client:
        raise HTTPException(404, "Mükellef bulunamadı")
    if not title.strip():
        raise HTTPException(400, "Belge adı zorunludur")
    orig = os.path.basename(file.filename or "belge")
    ext = os.path.splitext(orig)[1].lower()
    if ext not in ALLOWED_DOC_EXT:
        raise HTTPException(400, f"Desteklenmeyen dosya türü: {ext or 'bilinmiyor'}. İzin verilen: PDF, JPG, PNG, XLS(X), DOC(X)")
    content = await file.read()
    size = len(content)
    if size == 0:
        raise HTTPException(400, "Boş dosya yüklenemez")
    if size > MAX_UPLOAD_MB * 1024 * 1024:
        raise HTTPException(400, f"Dosya boyutu {MAX_UPLOAD_MB} MB sınırını aşıyor")
    # secure random physical name; client isolation via subfolder
    stored_name = f"{uuid.uuid4().hex}{ext}"
    client_dir = os.path.join(UPLOAD_DIR, cid)
    os.makedirs(client_dir, exist_ok=True)
    abs_path = os.path.join(client_dir, stored_name)
    with open(abs_path, "wb") as f:
        f.write(content)
    version = 1
    parent = None
    if parent_document_id:
        parent = await db.documents.find_one({"_id": ObjectId(parent_document_id), "client_id": cid})
        if parent:
            root_id = parent.get("parent_document_id") or parent_document_id
            cnt = await db.documents.count_documents({
                "client_id": cid, "deleted_at": None,
                "$or": [{"_id": ObjectId(root_id)}, {"parent_document_id": root_id}],
            })
            version = cnt + 1
            parent_document_id = root_id
    doc = {
        "client_id": cid, "title": title.strip(), "category": category or "Diğer",
        "subcategory": subcategory or "", "document_date": document_date or "", "period": period or "",
        "description": description or "", "tags": [t.strip() for t in tags.split(",") if t.strip()],
        "original_filename": orig, "stored_key": f"{cid}/{stored_name}",
        "mime_type": file.content_type or mimetypes.guess_type(orig)[0] or "application/octet-stream",
        "file_size": size, "expiry_date": expiry_date or "", "version": version,
        "parent_document_id": parent_document_id or None, "uploaded_by": user.get("name"),
        "created_at": now_utc().isoformat(), "updated_at": now_utc().isoformat(), "deleted_at": None,
    }
    r = await db.documents.insert_one(doc)
    await log_audit(user, "create", "document", cid, f"{title} ({orig})")
    doc["_id"] = r.inserted_id
    return _doc_serialize(doc)

@api.get("/clients/{cid}/documents")
async def list_documents(cid: str, user=Depends(get_current_user), category: Optional[str] = None,
                         period: Optional[str] = None, q: Optional[str] = None, tag: Optional[str] = None):
    query = {"client_id": cid, "deleted_at": None}
    if category:
        query["category"] = category
    if period:
        query["period"] = period
    if tag:
        query["tags"] = tag
    if q:
        query["title"] = {"$regex": q, "$options": "i"}
    docs = await db.documents.find(query).sort("created_at", -1).to_list(2000)
    return [_doc_serialize(d) for d in docs]

@api.get("/clients/{cid}/documents/{did}")
async def get_document(cid: str, did: str, user=Depends(get_current_user)):
    d = await db.documents.find_one({"_id": ObjectId(did), "client_id": cid, "deleted_at": None})
    if not d:
        raise HTTPException(404, "Belge bulunamadı")
    return _doc_serialize(d)

@api.get("/clients/{cid}/documents/{did}/download")
async def download_document(cid: str, did: str, inline: Optional[str] = None, user=Depends(get_current_user)):
    d = await db.documents.find_one({"_id": ObjectId(did), "client_id": cid, "deleted_at": None})
    if not d:
        raise HTTPException(404, "Belge bulunamadı")
    # stored_key is always relative "cid/name" — rebuild safely, no traversal
    safe_rel = os.path.normpath(d["stored_key"]).replace("\\", "/")
    if safe_rel.startswith("..") or not safe_rel.startswith(f"{cid}/"):
        raise HTTPException(400, "Geçersiz dosya yolu")
    abs_path = os.path.join(UPLOAD_DIR, safe_rel)
    if not os.path.isfile(abs_path):
        raise HTTPException(404, "Dosya fiziksel olarak bulunamadı")
    from urllib.parse import quote
    disp = "inline" if inline else "attachment"
    fn = quote(d.get("original_filename", "belge"))
    return FileResponse(abs_path, media_type=d.get("mime_type", "application/octet-stream"),
                        headers={"Content-Disposition": f"{disp}; filename*=UTF-8''{fn}"})

@api.put("/clients/{cid}/documents/{did}")
async def update_document(cid: str, did: str, body: dict, user=Depends(get_current_user)):
    d = await db.documents.find_one({"_id": ObjectId(did), "client_id": cid, "deleted_at": None})
    if not d:
        raise HTTPException(404, "Belge bulunamadı")
    allowed = {"title", "category", "subcategory", "document_date", "period", "description", "tags", "expiry_date"}
    upd = {k: v for k, v in body.items() if k in allowed}
    if "title" in upd and not str(upd["title"]).strip():
        raise HTTPException(400, "Belge adı zorunludur")
    if "tags" in upd and isinstance(upd["tags"], str):
        upd["tags"] = [t.strip() for t in upd["tags"].split(",") if t.strip()]
    upd["updated_at"] = now_utc().isoformat()
    await db.documents.update_one({"_id": ObjectId(did)}, {"$set": upd})
    await log_audit(user, "update", "document", cid, d.get("title"))
    nd = await db.documents.find_one({"_id": ObjectId(did)})
    return _doc_serialize(nd)

@api.delete("/clients/{cid}/documents/{did}")
async def delete_document(cid: str, did: str, user=Depends(get_current_user)):
    d = await db.documents.find_one({"_id": ObjectId(did), "client_id": cid, "deleted_at": None})
    if not d:
        raise HTTPException(404, "Belge bulunamadı")
    await db.documents.update_one({"_id": ObjectId(did)}, {"$set": {"deleted_at": now_utc().isoformat()}})
    await log_audit(user, "delete", "document", cid, d.get("title"))
    return {"ok": True}

@api.post("/accrual/run")
async def run_accrual(body: dict, user=Depends(get_current_user)):
    period = body["period"]  # YYYY-MM
    clients = await db.clients.find({"aktif": True}).to_list(2000)
    count = 0
    for c in clients:
        fee = c.get("aylik_ucret") or 0
        if fee <= 0:
            continue
        cid = str(c["_id"])
        exists = await db.transactions.find_one({"client_id": cid, "type": "borc", "accrual_period": period})
        if exists:
            continue
        await db.transactions.insert_one({"client_id": cid, "type": "borc", "amount": fee,
            "aciklama": f"{period} Mali Müşavirlik Hizmet Bedeli", "accrual_period": period,
            "date": f"{period}-01", "created_at": now_utc().isoformat()})
        count += 1
    await log_audit(user, "accrual", "transaction", None, f"{period} ({count})")
    return {"ok": True, "count": count}

# ---------------- tasks ----------------
@api.get("/tasks")
async def list_tasks(user=Depends(get_current_user), status: Optional[str] = None, client_id: Optional[str] = None):
    query = {}
    if status:
        query["status"] = status
    if client_id:
        query["client_id"] = client_id
    docs = await db.tasks.find(query).sort("deadline", 1).to_list(2000)
    out = []
    for d in docs:
        s = serialize(d)
        if s.get("client_id"):
            c = await db.clients.find_one({"_id": ObjectId(s["client_id"])})
            s["client_name"] = c["unvan"] if c else None
        out.append(s)
    return out

@api.post("/tasks")
async def create_task(body: dict, user=Depends(get_current_user)):
    body["created_at"] = now_utc().isoformat()
    body.setdefault("status", "Bekliyor")
    body.setdefault("olusturan", user.get("name"))
    r = await db.tasks.insert_one(body)
    await log_audit(user, "create", "task", str(r.inserted_id), body.get("baslik"))
    body["_id"] = r.inserted_id
    return serialize(body)

@api.put("/tasks/{tid}")
async def update_task(tid: str, body: dict, user=Depends(get_current_user)):
    body.pop("id", None); body.pop("_id", None); body.pop("client_name", None)
    await db.tasks.update_one({"_id": ObjectId(tid)}, {"$set": body})
    d = await db.tasks.find_one({"_id": ObjectId(tid)})
    return serialize(d)

@api.delete("/tasks/{tid}")
async def delete_task(tid: str, user=Depends(get_current_user)):
    await db.tasks.delete_one({"_id": ObjectId(tid)})
    return {"ok": True}

# ---------------- calendar ----------------
@api.get("/calendar")
async def calendar_events(user=Depends(get_current_user)):
    docs = await db.calendar_events.find().sort("date", 1).to_list(2000)
    return [serialize(d) for d in docs]

@api.post("/calendar")
async def create_event(body: dict, user=Depends(get_current_user)):
    body["created_at"] = now_utc().isoformat()
    r = await db.calendar_events.insert_one(body)
    body["_id"] = r.inserted_id
    return serialize(body)

@api.delete("/calendar/{eid}")
async def delete_event(eid: str, user=Depends(get_current_user)):
    await db.calendar_events.delete_one({"_id": ObjectId(eid)})
    return {"ok": True}

# ---------------- notifications ----------------
@api.get("/notifications")
async def get_notifications(user=Depends(get_current_user)):
    period = now_utc().strftime("%Y-%m")
    notes = []
    today = now_utc().date()
    # overdue/upcoming declarations
    decls = await db.declarations.find({"period": period}).to_list(20000)
    clients_cache = {}
    async def cname(cid):
        if cid not in clients_cache:
            c = await db.clients.find_one({"_id": ObjectId(cid)})
            clients_cache[cid] = c["unvan"] if c else "?"
        return clients_cache[cid]
    active = await db.clients.find({"aktif": True}).to_list(2000)
    for c in active:
        cid = str(c["_id"])
        for t in c.get("beyanname_turleri", []):
            d = next((x for x in decls if x["client_id"] == cid and x["type"] == t), None)
            status = d["status"] if d else "Hazırlanmadı"
            if status in ("Tamamlandı", "Gönderildi", "Tahakkuk Alındı", "Muaf"):
                continue
            dl = decl_deadline(t, period)
            days = (date.fromisoformat(dl) - today).days
            if days < 0:
                notes.append({"level": "kritik", "text": f"{c['unvan']} — {t} beyannamesi gecikti ({-days} gün)", "type": "beyanname"})
            elif days <= 3:
                notes.append({"level": "uyari", "text": f"{c['unvan']} — {t} son tarihine {days} gün kaldı", "type": "beyanname"})
    # overdue receivables
    txns = await db.transactions.find().to_list(20000)
    bal = {}
    for t in txns:
        bal[t["client_id"]] = bal.get(t["client_id"], 0) + (t["amount"] if t["type"] == "borc" else -t["amount"])
    for cid, b in bal.items():
        if b > 0:
            nm = await cname(cid)
            if b >= 10000:
                notes.append({"level": "uyari", "text": f"{nm} — {b:,.0f} TL ödenmemiş bakiye", "type": "cari"})
    notes.sort(key=lambda n: {"kritik": 0, "uyari": 1, "bilgi": 2}[n["level"]])
    return notes

# ---------------- dashboard ----------------
@api.get("/dashboard")
async def dashboard(user=Depends(get_current_user)):
    period = now_utc().strftime("%Y-%m")
    today = now_utc().date()
    active = await db.clients.find({"aktif": True}).to_list(2000)
    total = len(active)
    by_type = {}
    for c in active:
        st = c.get("sirket_turu", "Diğer")
        by_type[st] = by_type.get(st, 0) + 1
    # declarations counts by type due this month
    decls = await db.declarations.find({"period": period}).to_list(20000)
    decl_status = {}
    pending_by_type = {}
    for c in active:
        cid = str(c["_id"])
        for t in c.get("beyanname_turleri", []):
            d = next((x for x in decls if x["client_id"] == cid and x["type"] == t), None)
            status = d["status"] if d else "Hazırlanmadı"
            decl_status[status] = decl_status.get(status, 0) + 1
            if status not in ("Tamamlandı", "Gönderildi", "Tahakkuk Alındı", "Muaf"):
                pending_by_type[t] = pending_by_type.get(t, 0) + 1
    # e-defter incomplete
    edefter_clients = [c for c in active if c.get("edefter")]
    erecs = await db.edefter.find({"period": period}).to_list(5000)
    erec_map = {r["client_id"]: r for r in erecs}
    edefter_eksik = 0
    for c in edefter_clients:
        r = erec_map.get(str(c["_id"]))
        steps = r.get("steps", {}) if r else {}
        if not all(steps.get(s) for s in EDEFTER_STEPS):
            edefter_eksik += 1
    # cari
    txns = await db.transactions.find().to_list(20000)
    bal = {}
    for t in txns:
        bal[t["client_id"]] = bal.get(t["client_id"], 0) + (t["amount"] if t["type"] == "borc" else -t["amount"])
    toplam_alacak = sum(b for b in bal.values() if b > 0)
    this_month_collected = sum(t["amount"] for t in txns if t["type"] == "alacak" and str(t.get("date", "")).startswith(period))
    this_month_accrued = sum(t["amount"] for t in txns if t["type"] == "borc" and str(t.get("date", "")).startswith(period))
    top_debtors = sorted(bal.items(), key=lambda x: -x[1])[:5]
    debtors_out = []
    for cid, b in top_debtors:
        if b <= 0:
            continue
        c = await db.clients.find_one({"_id": ObjectId(cid)})
        debtors_out.append({"unvan": c["unvan"] if c else "?", "client_id": cid, "bakiye": b})
    # tasks
    open_tasks = await db.tasks.count_documents({"status": {"$nin": ["Tamamlandı", "İptal"]}})
    overdue_tasks = await db.tasks.count_documents({"status": {"$nin": ["Tamamlandı", "İptal"]}, "deadline": {"$lt": today.isoformat()}})
    # upcoming 7 days deadlines
    upcoming = []
    for c in active:
        cid = str(c["_id"])
        for t in c.get("beyanname_turleri", []):
            d = next((x for x in decls if x["client_id"] == cid and x["type"] == t), None)
            status = d["status"] if d else "Hazırlanmadı"
            if status in ("Tamamlandı", "Gönderildi", "Tahakkuk Alındı", "Muaf"):
                continue
            dl = decl_deadline(t, period)
            days = (date.fromisoformat(dl) - today).days
            if -30 <= days <= 7:
                upcoming.append({"unvan": c["unvan"], "type": t, "deadline": dl, "days": days, "client_id": cid})
    upcoming.sort(key=lambda x: x["days"])
    return {
        "period": period,
        "total_clients": total,
        "by_type": by_type,
        "pending_by_type": pending_by_type,
        "decl_status": decl_status,
        "edefter_eksik": edefter_eksik,
        "edefter_total": len(edefter_clients),
        "toplam_alacak": toplam_alacak,
        "this_month_collected": this_month_collected,
        "this_month_accrued": this_month_accrued,
        "top_debtors": debtors_out,
        "open_tasks": open_tasks,
        "overdue_tasks": overdue_tasks,
        "upcoming": upcoming[:15],
    }

# ---------------- audit ----------------
@api.get("/audit")
async def get_audit(user=Depends(get_current_user)):
    if user.get("role") not in ("admin", "mali_musavir"):
        raise HTTPException(403, "Yetkiniz yok")
    docs = await db.audit_logs.find().sort("created_at", -1).to_list(300)
    return [serialize(d) for d in docs]

# ---------------- reports ----------------
@api.get("/reports")
async def reports(user=Depends(get_current_user)):
    active = await db.clients.find({"aktif": True}).to_list(2000)
    period = now_utc().strftime("%Y-%m")
    decls = await db.declarations.find({"period": period}).to_list(20000)
    total_decl = 0; done_decl = 0
    for c in active:
        for t in c.get("beyanname_turleri", []):
            total_decl += 1
            d = next((x for x in decls if x["client_id"] == str(c["_id"]) and x["type"] == t), None)
            if d and d["status"] in ("Tamamlandı", "Gönderildi", "Tahakkuk Alındı"):
                done_decl += 1
    # personel workload
    workload = {}
    for c in active:
        p = c.get("sorumlu_personel", "Atanmamış")
        workload[p] = workload.get(p, 0) + 1
    txns = await db.transactions.find().to_list(20000)
    monthly = {}
    for t in txns:
        m = str(t.get("date", ""))[:7]
        if not m:
            continue
        monthly.setdefault(m, {"tahakkuk": 0, "tahsilat": 0})
        if t["type"] == "borc":
            monthly[m]["tahakkuk"] += t["amount"]
        else:
            monthly[m]["tahsilat"] += t["amount"]
    monthly_list = [{"month": k, **v} for k, v in sorted(monthly.items())][-6:]
    return {
        "total_clients": len(active),
        "declaration_completion": round(done_decl / total_decl * 100) if total_decl else 0,
        "workload": [{"personel": k, "adet": v} for k, v in sorted(workload.items(), key=lambda x: -x[1])],
        "monthly": monthly_list,
    }

# ---------------- AI assistant ----------------
@api.post("/assistant")
async def assistant(body: dict, user=Depends(get_current_user)):
    question = body.get("question", "")
    period = now_utc().strftime("%Y-%m")
    active = await db.clients.find({"aktif": True}).to_list(2000)
    decls = await db.declarations.find({"period": period}).to_list(20000)
    txns = await db.transactions.find().to_list(20000)
    tasks = await db.tasks.find({"status": {"$nin": ["Tamamlandı", "İptal"]}}).to_list(2000)
    bal = {}
    for t in txns:
        bal[t["client_id"]] = bal.get(t["client_id"], 0) + (t["amount"] if t["type"] == "borc" else -t["amount"])
    # build compact context
    lines = [f"Dönem: {period}. Toplam aktif mükellef: {len(active)}."]
    for c in active[:60]:
        cid = str(c["_id"])
        pend = []
        for t in c.get("beyanname_turleri", []):
            d = next((x for x in decls if x["client_id"] == cid and x["type"] == t), None)
            s = d["status"] if d else "Hazırlanmadı"
            if s not in ("Tamamlandı", "Gönderildi", "Tahakkuk Alındı", "Muaf"):
                pend.append(t)
        b = bal.get(cid, 0)
        lines.append(f"- {c['unvan']} ({c.get('sirket_turu','')}, sorumlu:{c.get('sorumlu_personel','-')}) | Bekleyen beyanname: {','.join(pend) or 'yok'} | Bakiye: {b:.0f} TL")
    for tk in tasks[:40]:
        lines.append(f"- GÖREV: {tk.get('baslik')} | sorumlu:{tk.get('sorumlu','-')} | son:{tk.get('deadline','-')} | durum:{tk.get('status')}")
    context = "\n".join(lines)
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        chat = LlmChat(api_key=os.environ["EMERGENT_LLM_KEY"], session_id=f"asst-{user['_id']}",
                       system_message=("Sen bir Türk mali müşavirlik ofisi için akıllı ofis asistanısın. "
                                       "Sadece aşağıda verilen ofis verilerine dayanarak Türkçe, kısa ve net cevap ver. "
                                       "Liste istenirse madde madde yaz. Veride yoksa 'Bu bilgi mevcut verilerde yok' de.\n\n"
                                       f"OFİS VERİLERİ:\n{context}")).with_model("openai", "gpt-4o-mini")
        resp = await chat.send_message(UserMessage(text=question))
        return {"answer": resp}
    except Exception as e:
        logger.error(f"assistant error: {e}")
        return {"answer": "Asistan şu anda yanıt veremedi. Lütfen tekrar deneyin."}

# ---------------- seeding ----------------
async def seed():
    admin_email = os.environ["ADMIN_EMAIL"].lower()
    admin_pw = os.environ["ADMIN_PASSWORD"]
    existing = await db.users.find_one({"email": admin_email})
    if not existing:
        await db.users.insert_one({"email": admin_email, "password_hash": hash_password(admin_pw),
                                   "name": "Suat Alptekin (SMMM)", "role": "admin", "created_at": now_utc().isoformat()})
    elif not verify_password(admin_pw, existing["password_hash"]):
        await db.users.update_one({"email": admin_email}, {"$set": {"password_hash": hash_password(admin_pw)}})
    # demo personel
    for em, nm, rl in [("ayse@ofis.com", "Ayşe Demir", "muhasebe"), ("mehmet@ofis.com", "Mehmet Kaya", "kidemli"), ("zeynep@ofis.com", "Zeynep Şahin", "bordro")]:
        if not await db.users.find_one({"email": em}):
            await db.users.insert_one({"email": em, "password_hash": hash_password("Ofis2026!"), "name": nm, "role": rl, "created_at": now_utc().isoformat()})

    if await db.clients.count_documents({}) > 0:
        return
    demo = [
        {"unvan": "Alptekin İnşaat Ltd. Şti.", "vkn": "1234567890", "vergi_dairesi": "Kadıköy", "sirket_turu": "Limited", "nace": "41.20.01", "faaliyet": "Bina inşaatı", "telefon": "0532 111 22 33", "email": "info@alptekin.com", "yetkili": "Osman Alptekin", "aylik_ucret": 8000, "edefter": True, "efatura": True, "beyanname_turleri": ["KDV1", "MUHSGK", "Geçici Vergi", "Damga"], "sorumlu_personel": "Ayşe Demir", "calisan_sayisi": 12, "aktif": True},
        {"unvan": "Kara Ticaret A.Ş.", "vkn": "2345678901", "vergi_dairesi": "Beşiktaş", "sirket_turu": "Anonim", "nace": "46.90.01", "faaliyet": "Toptan ticaret", "telefon": "0533 222 33 44", "email": "muhasebe@karaticaret.com", "yetkili": "Deniz Kara", "aylik_ucret": 15000, "edefter": True, "efatura": True, "beyanname_turleri": ["KDV1", "KDV2", "MUHSGK", "Geçici Vergi", "Kurumlar", "Damga"], "sorumlu_personel": "Mehmet Kaya", "calisan_sayisi": 34, "aktif": True},
        {"unvan": "Deniz Danışmanlık", "vkn": "3456789012", "tckn": "12345678901", "vergi_dairesi": "Şişli", "sirket_turu": "Şahıs", "nace": "70.22.01", "faaliyet": "İşletme danışmanlığı", "telefon": "0534 333 44 55", "email": "deniz@danismanlik.com", "yetkili": "Deniz Yıldız", "aylik_ucret": 4500, "edefter": False, "efatura": True, "beyanname_turleri": ["KDV1", "MUHSGK", "Geçici Vergi", "Gelir Vergisi"], "sorumlu_personel": "Ayşe Demir", "calisan_sayisi": 3, "aktif": True},
        {"unvan": "Yılmaz Otomotiv Ltd. Şti.", "vkn": "4567890123", "vergi_dairesi": "Ümraniye", "sirket_turu": "Limited", "nace": "45.20.01", "faaliyet": "Motorlu taşıt bakım onarım", "telefon": "0535 444 55 66", "email": "info@yilmazoto.com", "yetkili": "Ahmet Yılmaz", "aylik_ucret": 6500, "edefter": True, "efatura": True, "beyanname_turleri": ["KDV1", "MUHSGK", "Geçici Vergi", "Damga"], "sorumlu_personel": "Zeynep Şahin", "calisan_sayisi": 8, "aktif": True},
        {"unvan": "Güneş Tekstil A.Ş.", "vkn": "5678901234", "vergi_dairesi": "Bağcılar", "sirket_turu": "Anonim", "nace": "13.10.01", "faaliyet": "Tekstil üretimi", "telefon": "0536 555 66 77", "email": "info@gunestekstil.com", "yetkili": "Fatma Güneş", "aylik_ucret": 12000, "edefter": True, "efatura": True, "beyanname_turleri": ["KDV1", "KDV2", "MUHSGK", "Geçici Vergi", "Kurumlar"], "sorumlu_personel": "Mehmet Kaya", "calisan_sayisi": 45, "aktif": True},
        {"unvan": "Aydın Market Ltd. Şti.", "vkn": "6789012345", "vergi_dairesi": "Maltepe", "sirket_turu": "Limited", "nace": "47.11.01", "faaliyet": "Market işletmeciliği", "telefon": "0537 666 77 88", "email": "info@aydinmarket.com", "yetkili": "Kemal Aydın", "aylik_ucret": 5500, "edefter": False, "efatura": True, "beyanname_turleri": ["KDV1", "MUHSGK", "Geçici Vergi"], "sorumlu_personel": "Ayşe Demir", "calisan_sayisi": 6, "aktif": True},
        {"unvan": "Öztürk Nakliyat", "vkn": "7890123456", "tckn": "23456789012", "vergi_dairesi": "Pendik", "sirket_turu": "Şahıs", "nace": "49.41.01", "faaliyet": "Karayolu taşımacılığı", "telefon": "0538 777 88 99", "email": "ozturk@nakliyat.com", "yetkili": "Hasan Öztürk", "aylik_ucret": 3800, "edefter": False, "efatura": False, "beyanname_turleri": ["KDV1", "MUHSGK", "Gelir Vergisi"], "sorumlu_personel": "Zeynep Şahin", "calisan_sayisi": 4, "aktif": True},
        {"unvan": "Çelik Yazılım A.Ş.", "vkn": "8901234567", "vergi_dairesi": "Ataşehir", "sirket_turu": "Anonim", "nace": "62.01.01", "faaliyet": "Yazılım geliştirme", "telefon": "0539 888 99 00", "email": "info@celikyazilim.com", "yetkili": "Burak Çelik", "aylik_ucret": 18000, "edefter": True, "efatura": True, "beyanname_turleri": ["KDV1", "MUHSGK", "Geçici Vergi", "Kurumlar", "Damga"], "sorumlu_personel": "Mehmet Kaya", "calisan_sayisi": 22, "aktif": True},
    ]
    res = await db.clients.insert_many(demo)
    ids = res.inserted_ids
    period = now_utc().strftime("%Y-%m")
    # some declarations & transactions & tasks
    statuses = ["Tamamlandı", "Gönderildi", "Hazırlanıyor", "Hazırlanmadı", "Kontrol Bekliyor"]
    for i, cid in enumerate(ids):
        c = demo[i]
        for j, t in enumerate(c["beyanname_turleri"]):
            await db.declarations.insert_one({"client_id": str(cid), "type": t, "period": period,
                "status": statuses[(i + j) % len(statuses)], "deadline": decl_deadline(t, period),
                "checklist": {}, "updated_at": now_utc().isoformat()})
        # accrual for this + last month
        for pm in [period, (now_utc().replace(day=1) - timedelta(days=1)).strftime("%Y-%m")]:
            await db.transactions.insert_one({"client_id": str(cid), "type": "borc", "amount": c["aylik_ucret"],
                "aciklama": f"{pm} Mali Müşavirlik Hizmet Bedeli", "accrual_period": pm, "date": f"{pm}-01", "created_at": now_utc().isoformat()})
        # partial payment for some
        if i % 2 == 0:
            await db.transactions.insert_one({"client_id": str(cid), "type": "alacak", "amount": c["aylik_ucret"],
                "aciklama": "Tahsilat - Havale", "yontem": "Havale", "date": now_utc().date().isoformat(), "created_at": now_utc().isoformat()})
        if c.get("edefter"):
            await db.edefter.insert_one({"client_id": str(cid), "period": period,
                "steps": {"kayitlar": i % 2 == 0, "kontrol": i % 3 == 0, "berat_olustur": False, "berat_yukle": False, "berat_onay": False, "arsiv": False},
                "updated_at": now_utc().isoformat()})
    # tasks
    demo_tasks = [
        {"baslik": "Alptekin İnşaat - Banka ekstresi iste", "client_id": str(ids[0]), "sorumlu": "Ayşe Demir", "oncelik": "Yüksek", "deadline": (now_utc().date() + timedelta(days=2)).isoformat(), "status": "Bekliyor"},
        {"baslik": "Kara Ticaret - Kurumlar geçici vergi kontrol", "client_id": str(ids[1]), "sorumlu": "Mehmet Kaya", "oncelik": "Orta", "deadline": (now_utc().date() + timedelta(days=5)).isoformat(), "status": "Devam Ediyor"},
        {"baslik": "Yılmaz Otomotiv - SGK işe giriş bildirgesi", "client_id": str(ids[3]), "sorumlu": "Zeynep Şahin", "oncelik": "Yüksek", "deadline": (now_utc().date() - timedelta(days=1)).isoformat(), "status": "Bekliyor"},
    ]
    for t in demo_tasks:
        t["created_at"] = now_utc().isoformat()
    await db.tasks.insert_many(demo_tasks)
    logger.info("Demo data seeded")

@app.on_event("startup")
async def startup():
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    await db.users.create_index("email", unique=True)
    await seed()
    # write test credentials
    try:
        Path("/app/memory/test_credentials.md").write_text(
            f"# Test Credentials\n\n## Admin (SMMM Owner)\n- Email: {os.environ['ADMIN_EMAIL']}\n- Password: {os.environ['ADMIN_PASSWORD']}\n- Role: admin\n\n"
            "## Demo Personel\n- ayse@ofis.com / Ofis2026! (muhasebe)\n- mehmet@ofis.com / Ofis2026! (kidemli)\n- zeynep@ofis.com / Ofis2026! (bordro)\n\n"
            "## Auth endpoints\n- POST /api/auth/login\n- POST /api/auth/logout\n- GET /api/auth/me\n"
        )
    except Exception as e:
        logger.error(f"cred write: {e}")

app.include_router(api)
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=[os.environ.get("FRONTEND_URL", "http://localhost:3000"), "http://localhost:3000"],
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown():
    client.close()
